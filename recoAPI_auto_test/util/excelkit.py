#!/usr/bin/env python3
# -*- coding: utf-8 -*- 
# @Time : 2019-05-14 21:27 
# @Author : lmh
# @Software: PyCharm
#

import openpyxl

class exceldata():
    def __init__(self, excelfile, sheetname):
        self.excelfile = excelfile
        self.sheetname = sheetname
        self.wb = openpyxl.load_workbook(self.excelfile)
        self.ws = self.wb[self.sheetname]
        self.rows = self.ws.max_row
        self.cols = self.ws.max_column

    def exceldata_row(self):
        """
        读取excel,每行数据和第一行组成一个字典.返回excel数据列表
        :return:excel数据列表,每个元素是该行和首行对应的字典
        """
        exceldata = []
        # 读取
        for r in range(1, self.rows+1):
            excelline = []
            for c in range(1, self.cols + 1):
                excelline.append(self.ws.cell(row=r, column=c).value)
            exceldata.append(excelline)

        excellist = []
        for i in range(1, len(exceldata)):
            dict1 = dict(zip(exceldata[0], exceldata[i]))
            excellist.append(dict1)
        return excellist

    def getChars(self, length):
        return [self.getChar(index) for index in range(length)]

    def getChar(self, number):
        factor, moder = divmod(number, 26) # 26 字母个数
        modChar = chr(moder + 65)          # 65 -> 'A'
        if factor != 0:
            modChar = self.getChar(factor-1) + modChar # factor - 1 : 商为有效值时起始数为 1 而余数是 0
        return modChar

    def exceldata_col(self):
        """
        读取excel,每列数据和第一列组成一个字典.返回excel数据列表
        :return:excel数据列表,每个元素是该行和首行对应的字典
        """
        # 读取
        exceldata = []
        for c in range(1, self.cols+1):
            excelline = []
            for r in range(1, self.rows + 1):
                excelline.append(self.ws.cell(row=r, column=c).value)
            exceldata.append(excelline)

        excellist = []
        for i in range(1, len(exceldata)):
            dict1 = dict(zip(exceldata[0], exceldata[i]))
            excellist.append(dict1)
        return excellist